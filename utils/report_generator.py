"""
report_generator.py — Excel and PDF export logic.

Both functions write to an in-memory BytesIO buffer so Streamlit
can serve them via st.download_button without touching the filesystem.
"""

import io
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from fpdf import FPDF


# ------------------------------------------------------------------
# Excel report
# ------------------------------------------------------------------

def generate_excel_report(
    df: pd.DataFrame,
    kpis: dict,
    platform_agg: pd.DataFrame,
) -> io.BytesIO:
    """
    Create a multi-sheet Excel workbook in memory.

    Sheet 1 — "Summary": KPI totals + platform-wise breakdown table.
    Sheet 2 — "Campaign Data": full dataset with computed metrics.

    Parameters
    ----------
    df : pd.DataFrame
        Cleaned DataFrame with metric columns.
    kpis : dict
        Output of metrics.compute_summary_kpis().
    platform_agg : pd.DataFrame
        Platform-level aggregated data.

    Returns
    -------
    io.BytesIO
        Buffer containing the .xlsx file.
    """
    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # ---- Sheet 1: Summary ----
        _write_summary_sheet(writer, kpis, platform_agg)

        # ---- Sheet 2: Campaign Data ----
        # Prepare a display-friendly copy
        display_df = df.copy()
        if "date" in display_df.columns:
            display_df["date"] = display_df["date"].dt.strftime("%Y-%m-%d")
        display_df.columns = [c.title() for c in display_df.columns]
        display_df.to_excel(writer, sheet_name="Campaign Data", index=False)

        # Style the Campaign Data sheet
        ws = writer.sheets["Campaign Data"]
        _style_header_row(ws)
        _auto_width(ws)

    buffer.seek(0)
    return buffer


def _write_summary_sheet(writer, kpis, platform_agg):
    """Write KPI summary and platform breakdown to the Summary sheet."""
    # Build a small summary DataFrame for the KPIs
    summary_data = {
        "Metric": [
            "Total Spend",
            "Total Revenue",
            "Total Clicks",
            "Total Impressions",
            "Total Conversions",
            "Overall CTR (%)",
            "Overall CPC",
            "Overall CPM",
            "Overall Conversion Rate (%)",
            "Overall ROI (%)",
        ],
        "Value": [
            f"₹{kpis['total_spend']:,.2f}",
            f"₹{kpis['total_revenue']:,.2f}",
            f"{kpis['total_clicks']:,.0f}",
            f"{kpis['total_impressions']:,.0f}",
            f"{kpis['total_conversions']:,.0f}",
            f"{kpis['overall_ctr']:.2f}%",
            f"₹{kpis['overall_cpc']:.2f}",
            f"₹{kpis['overall_cpm']:.2f}",
            f"{kpis['overall_conv_rate']:.2f}%",
            f"{kpis['overall_roi']:.2f}%",
        ],
    }
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name="Summary", index=False, startrow=0)

    # Platform breakdown below the KPI block
    start_row = len(summary_df) + 3  # leave a gap
    platform_display = platform_agg.copy()
    platform_display.columns = [c.title() for c in platform_display.columns]
    platform_display.to_excel(
        writer, sheet_name="Summary", index=False, startrow=start_row
    )

    ws = writer.sheets["Summary"]
    _style_header_row(ws)
    # Also style the platform table header
    header_row = start_row + 1  # openpyxl is 1-indexed
    for col_idx in range(1, len(platform_display.columns) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E4057", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    _auto_width(ws)


def _style_header_row(ws):
    """Apply bold + dark background to the first row of a worksheet."""
    header_fill = PatternFill(start_color="2E4057", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws):
    """Auto-fit column widths based on content length."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


# ------------------------------------------------------------------
# PDF report
# ------------------------------------------------------------------

def generate_pdf_report(
    kpis: dict,
    campaign_agg: pd.DataFrame,
    date_range: tuple = None,
) -> io.BytesIO:
    """
    Generate a simple one-page PDF report in memory.

    Content: title, date range, KPI summary block,
    top 5 campaigns table, bottom 5 campaigns table.

    Parameters
    ----------
    kpis : dict
        Output of metrics.compute_summary_kpis().
    campaign_agg : pd.DataFrame
        Campaign-level aggregated data (must include roi (%) column).
    date_range : tuple of (min_date, max_date), optional
        Date boundaries for the report header.

    Returns
    -------
    io.BytesIO
        Buffer containing the PDF file.
    """
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # ---- Title ----
    pdf.set_font("Helvetica", "B", 18)
    pdf.cell(0, 12, "Campaign Performance Report", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(4)

    # ---- Date range ----
    if date_range:
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(
            0, 8,
            f"Period: {date_range[0].strftime('%d %b %Y')} - {date_range[1].strftime('%d %b %Y')}",
            new_x="LMARGIN", new_y="NEXT", align="C",
        )
        pdf.ln(4)

    # ---- KPI Summary ----
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 10, "Key Performance Indicators", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    kpi_rows = [
        ("Total Spend", f"Rs. {kpis['total_spend']:,.2f}"),
        ("Total Revenue", f"Rs. {kpis['total_revenue']:,.2f}"),
        ("Total Clicks", f"{kpis['total_clicks']:,.0f}"),
        ("Overall CTR", f"{kpis['overall_ctr']:.2f}%"),
        ("Overall CPC", f"Rs. {kpis['overall_cpc']:.2f}"),
        ("Overall ROI", f"{kpis['overall_roi']:.2f}%"),
    ]

    pdf.set_font("Helvetica", "", 11)
    for label, value in kpi_rows:
        pdf.cell(80, 8, label, border=1)
        pdf.cell(80, 8, value, border=1, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(6)

    # ---- Top 5 campaigns by ROI ----
    _add_campaign_table(pdf, campaign_agg, "Top 5 Campaigns by ROI", ascending=False)
    pdf.ln(6)

    # ---- Bottom 5 campaigns by ROI ----
    _add_campaign_table(pdf, campaign_agg, "Bottom 5 Campaigns by ROI", ascending=True)

    # Write to buffer
    buffer = io.BytesIO()
    pdf.output(buffer)
    buffer.seek(0)
    return buffer


def _add_campaign_table(
    pdf: FPDF,
    campaign_agg: pd.DataFrame,
    title: str,
    ascending: bool,
    n: int = 5,
):
    """Add a ranked campaign table section to the PDF."""
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 10, title, new_x="LMARGIN", new_y="NEXT")

    sorted_df = campaign_agg.sort_values("roi (%)", ascending=ascending).head(n)

    # Header row
    col_widths = [65, 30, 30, 30, 30]
    headers = ["Campaign", "Spend", "Revenue", "Clicks", "ROI (%)"]
    pdf.set_font("Helvetica", "B", 9)
    for w, h in zip(col_widths, headers):
        pdf.cell(w, 7, h, border=1, align="C")
    pdf.ln()

    # Data rows
    pdf.set_font("Helvetica", "", 9)
    for _, row in sorted_df.iterrows():
        name = str(row.get("campaign name", ""))[:28]  # truncate long names
        pdf.cell(col_widths[0], 7, name, border=1)
        pdf.cell(col_widths[1], 7, f"{row['spend']:,.0f}", border=1, align="R")
        pdf.cell(col_widths[2], 7, f"{row['revenue']:,.0f}", border=1, align="R")
        pdf.cell(col_widths[3], 7, f"{row['clicks']:,.0f}", border=1, align="R")
        pdf.cell(col_widths[4], 7, f"{row['roi (%)']:.1f}%", border=1, align="R")
        pdf.ln()
